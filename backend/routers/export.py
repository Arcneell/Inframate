"""
Export Router - CSV and Excel export functionality for data.
Provides export endpoints for equipment, tickets, contracts, and software.
"""
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import and_
from typing import Optional, List, Literal
from datetime import datetime, timezone
import csv
import io
import re
import html
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from backend.core.database import get_db
from backend.core.security import get_current_user, get_current_admin_user
from backend import models, schemas

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/export", tags=["export"])


def generate_csv(data: List[dict], filename: str) -> StreamingResponse:
    """Generate a CSV file from a list of dictionaries."""
    if not data:
        output = io.StringIO()
        output.write("No data to export")
        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=data[0].keys())
    writer.writeheader()
    writer.writerows(data)
    output.seek(0)

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


def strip_html(text: str) -> str:
    """Remove HTML tags and clean up text for CSV export."""
    if not text:
        return ""
    # Remove img tags completely (they contain base64 data)
    text = re.sub(r'<img[^>]*>', '', text)
    # Replace common block elements with newlines
    text = re.sub(r'</?(p|div|h[1-6]|li|tr|br)[^>]*>', '\n', text, flags=re.IGNORECASE)
    # Remove all remaining HTML tags
    text = re.sub(r'<[^>]+>', '', text)
    # Decode HTML entities
    text = html.unescape(text)
    # Clean up whitespace: multiple newlines to single, trim lines
    text = re.sub(r'\n\s*\n', '\n', text)
    text = '\n'.join(line.strip() for line in text.split('\n'))
    return text.strip()


def generate_xlsx(data: List[dict], filename: str, column_labels: dict = None) -> StreamingResponse:
    """Generate an Excel XLSX file from a list of dictionaries with formatting."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Export"

    if not data:
        ws['A1'] = "No data to export"
    else:
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_alignment = Alignment(vertical="top", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='E5E7EB'),
            right=Side(style='thin', color='E5E7EB'),
            top=Side(style='thin', color='E5E7EB'),
            bottom=Side(style='thin', color='E5E7EB')
        )
        alt_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

        # Get column keys
        columns = list(data[0].keys())

        # Write headers with labels if provided
        for col_idx, col_key in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = column_labels.get(col_key, col_key) if column_labels else col_key
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Write data rows
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, col_key in enumerate(columns, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = row_data.get(col_key, "")
                cell.alignment = cell_alignment
                cell.border = thin_border
                # Alternate row colors
                if row_idx % 2 == 0:
                    cell.fill = alt_fill

        # Auto-adjust column widths
        for col_idx, col_key in enumerate(columns, 1):
            max_length = len(column_labels.get(col_key, col_key) if column_labels else col_key)
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        # For multiline text, get the longest line
                        lines = str(cell.value).split('\n')
                        cell_max = max(len(line) for line in lines)
                        max_length = max(max_length, min(cell_max, 50))  # Cap at 50 chars
            ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

        # Freeze header row
        ws.freeze_panes = 'A2'

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ==================== EQUIPMENT EXPORT ====================

@router.get("/equipment")
def export_equipment(
    status: Optional[str] = None,
    equipment_type_id: Optional[int] = None,
    location_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_admin_user)
):
    """Export equipment list to CSV."""
    query = db.query(models.Equipment).options(
        joinedload(models.Equipment.equipment_type),
        joinedload(models.Equipment.model),
        joinedload(models.Equipment.location),
        joinedload(models.Equipment.supplier)
    )

    # Apply entity filter
    if current_user.entity_id:
        query = query.filter(models.Equipment.entity_id == current_user.entity_id)

    # Apply filters
    if status:
        query = query.filter(models.Equipment.status == status)
    if equipment_type_id:
        query = query.filter(models.Equipment.equipment_type_id == equipment_type_id)
    if location_id:
        query = query.filter(models.Equipment.location_id == location_id)

    equipment_list = query.order_by(models.Equipment.name).all()

    data = []
    for eq in equipment_list:
        data.append({
            "id": eq.id,
            "name": eq.name,
            "status": eq.status,
            "type": eq.equipment_type.name if eq.equipment_type else "",
            "model": eq.model.name if eq.model else "",
            "manufacturer": eq.model.manufacturer.name if eq.model and eq.model.manufacturer else "",
            "serial_number": eq.serial_number or "",
            "asset_tag": eq.asset_tag or "",
            "location": eq.location.name if eq.location else "",
            "supplier": eq.supplier.name if eq.supplier else "",
            "purchase_date": eq.purchase_date.isoformat() if eq.purchase_date else "",
            "warranty_expiry": eq.warranty_expiry.isoformat() if eq.warranty_expiry else "",
            "purchase_price": eq.purchase_price or "",
            "ip_address": eq.remote_ip or "",
            "rack_position": f"U{eq.position_u}" if eq.position_u else "",
            "notes": eq.notes or "",
            "created_at": eq.created_at.isoformat() if eq.created_at else ""
        })

    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"equipment_export_{timestamp}.csv"

    logger.info(f"Equipment export generated by {current_user.username}: {len(data)} items")
    return generate_csv(data, filename)


# ==================== TICKETS EXPORT ====================

@router.get("/tickets")
def export_tickets(
    status: Optional[str] = None,
    priority: Optional[str] = None,
    ticket_type: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    """Export tickets to CSV."""
    query = db.query(models.Ticket).options(
        joinedload(models.Ticket.requester),
        joinedload(models.Ticket.assigned_to),
        joinedload(models.Ticket.equipment)
    )

    # Access control - tech, admin, and superadmin can see all tickets in their entity
    if current_user.role not in ("tech", "admin", "superadmin"):
        query = query.filter(models.Ticket.requester_id == current_user.id)
    elif current_user.entity_id:
        query = query.filter(models.Ticket.entity_id == current_user.entity_id)

    # Apply filters
    if status:
        query = query.filter(models.Ticket.status == status)
    if priority:
        query = query.filter(models.Ticket.priority == priority)
    if ticket_type:
        query = query.filter(models.Ticket.ticket_type == ticket_type)
    if date_from:
        try:
            from_date = datetime.fromisoformat(date_from)
            query = query.filter(models.Ticket.created_at >= from_date)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date_from format. Use ISO format.")
    if date_to:
        try:
            to_date = datetime.fromisoformat(date_to)
            query = query.filter(models.Ticket.created_at <= to_date)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date_to format. Use ISO format.")

    tickets = query.order_by(models.Ticket.created_at.desc()).all()

    data = []
    for t in tickets:
        data.append({
            "ticket_number": t.ticket_number,
            "title": t.title,
            "description": strip_html(t.description) if t.description else "",
            "type": t.ticket_type,
            "category": t.category or "",
            "status": t.status,
            "priority": t.priority,
            "requester": t.requester.username if t.requester else "",
            "assigned_to": t.assigned_to.username if t.assigned_to else "",
            "equipment": t.equipment.name if t.equipment else "",
            "sla_due_date": t.sla_due_date.isoformat() if t.sla_due_date else "",
            "sla_breached": "Yes" if t.sla_breached else "No",
            "resolution": strip_html(t.resolution) if t.resolution else "",
            "created_at": t.created_at.isoformat() if t.created_at else "",
            "resolved_at": t.resolved_at.isoformat() if t.resolved_at else "",
            "closed_at": t.closed_at.isoformat() if t.closed_at else ""
        })

    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"tickets_export_{timestamp}.csv"

    logger.info(f"Tickets export generated by {current_user.username}: {len(data)} items")
    return generate_csv(data, filename)


@router.post("/tickets/bulk")
def export_tickets_bulk(
    request: schemas.BulkTicketExport,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    """Export selected tickets with specific columns to CSV."""
    # Define all available columns and their mapping functions
    # Note: description and resolution use strip_html to remove HTML formatting
    column_mapping = {
        "ticket_number": lambda t: t.ticket_number,
        "title": lambda t: t.title,
        "description": lambda t: strip_html(t.description) if t.description else "",
        "type": lambda t: t.ticket_type,
        "category": lambda t: t.category or "",
        "status": lambda t: t.status,
        "priority": lambda t: t.priority,
        "requester": lambda t: t.requester.username if t.requester else "",
        "requester_email": lambda t: t.requester.email if t.requester else "",
        "assigned_to": lambda t: t.assigned_to.username if t.assigned_to else "",
        "assigned_email": lambda t: t.assigned_to.email if t.assigned_to else "",
        "equipment": lambda t: t.equipment.name if t.equipment else "",
        "entity": lambda t: t.entity.name if t.entity else "",
        "sla_due_date": lambda t: t.sla_due_date.isoformat() if t.sla_due_date else "",
        "sla_breached": lambda t: "Yes" if t.sla_breached else "No",
        "resolution": lambda t: strip_html(t.resolution) if t.resolution else "",
        "created_at": lambda t: t.created_at.isoformat() if t.created_at else "",
        "updated_at": lambda t: t.updated_at.isoformat() if t.updated_at else "",
        "resolved_at": lambda t: t.resolved_at.isoformat() if t.resolved_at else "",
        "closed_at": lambda t: t.closed_at.isoformat() if t.closed_at else ""
    }

    # Validate columns
    invalid_columns = [col for col in request.columns if col not in column_mapping]
    if invalid_columns:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid columns: {', '.join(invalid_columns)}. Valid columns: {', '.join(column_mapping.keys())}"
        )

    # Fetch tickets by IDs
    query = db.query(models.Ticket).options(
        joinedload(models.Ticket.requester),
        joinedload(models.Ticket.assigned_to),
        joinedload(models.Ticket.equipment),
        joinedload(models.Ticket.entity)
    ).filter(models.Ticket.id.in_(request.ticket_ids))

    # Access control
    if current_user.role not in ("tech", "admin", "superadmin"):
        query = query.filter(models.Ticket.requester_id == current_user.id)
    elif current_user.entity_id:
        query = query.filter(models.Ticket.entity_id == current_user.entity_id)

    tickets = query.order_by(models.Ticket.created_at.desc()).all()

    if not tickets:
        raise HTTPException(status_code=404, detail="No tickets found matching the selection")

    # Build export data with only selected columns
    data = []
    for t in tickets:
        row = {}
        for col in request.columns:
            row[col] = column_mapping[col](t)
        data.append(row)

    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")

    # Column labels for nice headers
    column_labels = {
        "ticket_number": "Ticket #",
        "title": "Title",
        "description": "Description",
        "type": "Type",
        "category": "Category",
        "status": "Status",
        "priority": "Priority",
        "requester": "Requester",
        "requester_email": "Requester Email",
        "assigned_to": "Assigned To",
        "assigned_email": "Assigned Email",
        "equipment": "Equipment",
        "entity": "Entity",
        "sla_due_date": "SLA Due Date",
        "sla_breached": "SLA Breached",
        "resolution": "Resolution",
        "created_at": "Created At",
        "updated_at": "Updated At",
        "resolved_at": "Resolved At",
        "closed_at": "Closed At"
    }

    logger.info(f"Bulk tickets export generated by {current_user.username}: {len(data)} items, format: {request.format}")

    if request.format == "xlsx":
        filename = f"tickets_export_{timestamp}.xlsx"
        return generate_xlsx(data, filename, column_labels)
    else:
        filename = f"tickets_export_{timestamp}.csv"
        return generate_csv(data, filename)


# ==================== CONTRACTS EXPORT ====================

@router.get("/contracts")
def export_contracts(
    contract_type: Optional[str] = None,
    status: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_admin_user)
):
    """Export contracts to CSV."""
    query = db.query(models.Contract).options(
        joinedload(models.Contract.supplier)
    )

    # Apply entity filter
    if current_user.entity_id:
        query = query.filter(models.Contract.entity_id == current_user.entity_id)

    # Apply filters
    if contract_type:
        query = query.filter(models.Contract.contract_type == contract_type)
    if status:
        query = query.filter(models.Contract.status == status)

    contracts = query.order_by(models.Contract.end_date.desc()).all()

    data = []
    for c in contracts:
        data.append({
            "id": c.id,
            "name": c.name,
            "contract_number": c.contract_number or "",
            "type": c.contract_type,
            "status": c.status,
            "supplier": c.supplier.name if c.supplier else "",
            "start_date": c.start_date.isoformat() if c.start_date else "",
            "end_date": c.end_date.isoformat() if c.end_date else "",
            "annual_cost": c.annual_cost or "",
            "auto_renewal": "Yes" if c.auto_renewal else "No",
            "notice_period_days": c.notice_period_days or "",
            "support_level": c.support_level or "",
            "notes": c.notes or "",
            "created_at": c.created_at.isoformat() if c.created_at else ""
        })

    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"contracts_export_{timestamp}.csv"

    logger.info(f"Contracts export generated by {current_user.username}: {len(data)} items")
    return generate_csv(data, filename)


# ==================== SOFTWARE EXPORT ====================

@router.get("/software")
def export_software(
    category: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_admin_user)
):
    """Export software catalog with license information to CSV."""
    query = db.query(models.Software)

    # Apply entity filter
    if current_user.entity_id:
        query = query.filter(models.Software.entity_id == current_user.entity_id)

    # Apply filters
    if category:
        query = query.filter(models.Software.category == category)

    software_list = query.order_by(models.Software.name).all()

    data = []
    for s in software_list:
        # Count licenses and installations
        license_count = len(s.licenses) if s.licenses else 0
        total_seats = sum(l.quantity for l in s.licenses) if s.licenses else 0
        installation_count = len(s.installations) if s.installations else 0

        data.append({
            "id": s.id,
            "name": s.name,
            "version": s.version or "",
            "publisher": s.publisher or "",
            "category": s.category or "",
            "license_type": s.license_type or "",
            "total_licenses": license_count,
            "total_seats": total_seats,
            "installations": installation_count,
            "compliance": "OK" if installation_count <= total_seats else "Over-deployed",
            "website": s.website or "",
            "notes": s.notes or "",
            "created_at": s.created_at.isoformat() if s.created_at else ""
        })

    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"software_export_{timestamp}.csv"

    logger.info(f"Software export generated by {current_user.username}: {len(data)} items")
    return generate_csv(data, filename)


# ==================== IP ADDRESSES EXPORT ====================

@router.get("/ip-addresses")
def export_ip_addresses(
    subnet_id: Optional[int] = None,
    status: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_admin_user)
):
    """Export IP addresses to CSV."""
    query = db.query(models.IPAddress).options(
        joinedload(models.IPAddress.subnet),
        joinedload(models.IPAddress.equipment)
    )

    # Apply entity filter via subnet
    if current_user.entity_id:
        query = query.join(models.Subnet).filter(
            models.Subnet.entity_id == current_user.entity_id
        )

    # Apply filters
    if subnet_id:
        query = query.filter(models.IPAddress.subnet_id == subnet_id)
    if status:
        query = query.filter(models.IPAddress.status == status)

    ip_addresses = query.order_by(models.IPAddress.address).all()

    data = []
    for ip in ip_addresses:
        data.append({
            "address": ip.address,
            "subnet": ip.subnet.cidr if ip.subnet else "",
            "status": ip.status,
            "hostname": ip.hostname or "",
            "mac_address": ip.mac_address or "",
            "equipment": ip.equipment.name if ip.equipment else "",
            "description": ip.description or "",
            "last_seen": ip.last_seen.isoformat() if ip.last_seen else "",
            "created_at": ip.created_at.isoformat() if ip.created_at else ""
        })

    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"ip_addresses_export_{timestamp}.csv"

    logger.info(f"IP addresses export generated by {current_user.username}: {len(data)} items")
    return generate_csv(data, filename)


# ==================== AUDIT LOGS EXPORT ====================

@router.get("/audit-logs")
def export_audit_logs(
    action: Optional[str] = None,
    resource_type: Optional[str] = None,
    username: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    limit: int = Query(default=1000, le=10000),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_admin_user)
):
    """Export audit logs to CSV (admin only)."""
    query = db.query(models.AuditLog)

    # Apply filters
    if action:
        query = query.filter(models.AuditLog.action == action)
    if resource_type:
        query = query.filter(models.AuditLog.resource_type == resource_type)
    if username:
        query = query.filter(models.AuditLog.username.ilike(f"%{username}%"))
    if date_from:
        try:
            from_date = datetime.fromisoformat(date_from)
            query = query.filter(models.AuditLog.timestamp >= from_date)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date_from format. Use ISO format.")
    if date_to:
        try:
            to_date = datetime.fromisoformat(date_to)
            query = query.filter(models.AuditLog.timestamp <= to_date)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date_to format. Use ISO format.")

    logs = query.order_by(models.AuditLog.timestamp.desc()).limit(limit).all()

    data = []
    for log in logs:
        data.append({
            "timestamp": log.timestamp.isoformat() if log.timestamp else "",
            "username": log.username or "",
            "action": log.action,
            "resource_type": log.resource_type or "",
            "resource_id": log.resource_id or "",
            "ip_address": log.ip_address or "",
            "extra_data": str(log.extra_data) if log.extra_data else ""
        })

    timestamp_str = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"audit_logs_export_{timestamp_str}.csv"

    logger.info(f"Audit logs export generated by {current_user.username}: {len(data)} items")
    return generate_csv(data, filename)
